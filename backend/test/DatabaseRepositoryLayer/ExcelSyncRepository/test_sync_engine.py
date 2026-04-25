import json
import math
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook
from unittest.mock import patch

from DatabaseRepositoryLayer.ExcelSyncRepository.sync_engine import SyncEngine


# --------------------------------------------------
# Helpers
# --------------------------------------------------

_DEFAULT_CFG = {
    "central_file": "",            # filled per test
    "dependent_files_dir": "",     # filled per test
    "central_sheet_index": 0,
    "central_header_row": 1,
    "central_match_column": "Datenkategorie",
    "central_value_column": "Wert",
    "dependent_sheet_index": 1,
    "dependent_header_row": 5,
    "dependent_match_column": "Datenkategorie",
    "dependent_value_column": "Wert",
}


def _write_config(tmp_path: Path, overrides: dict | None = None) -> Path:
    cfg = {**_DEFAULT_CFG, **(overrides or {})}
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg), encoding="utf-8")
    return cfg_file


def _create_central_xlsx(path: Path, rows: list[tuple[str, object]]) -> None:
    """
    Create a minimal central xlsx with a single sheet.
    Row 1 = header (Datenkategorie | Wert), then data rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Datenkategorie", "Wert"])
    for cat, val in rows:
        ws.append([cat, val])
    wb.save(path)
    wb.close()


def _create_dependent_xlsx(
    path: Path,
    rows: list[tuple[str, object]],
    header_row: int = 5,
) -> None:
    """
    Create a dependent xlsx with two sheets.
    Sheet 0 is a dummy; sheet 1 has the data.
    Header row defaults to row 5.
    """
    wb = Workbook()
    wb.active.title = "Overview"
    ws = wb.create_sheet("Daten Generell")

    # Fill header at the specified row.
    ws.cell(row=header_row, column=1, value="Datenkategorie")
    ws.cell(row=header_row, column=2, value="Wert")

    for i, (cat, val) in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=val)

    wb.save(path)
    wb.close()


@pytest.fixture
def setup_dirs(tmp_path: Path):
    """Prepare central file dir and dependent dir."""
    central_dir = tmp_path / "central"
    central_dir.mkdir()
    dep_dir = tmp_path / "dependent"
    dep_dir.mkdir()
    return central_dir, dep_dir


# --------------------------------------------------
# __init__  /  config loading
# --------------------------------------------------

def test_init_loads_config(tmp_path: Path):
    cfg_file = _write_config(tmp_path, {
        "central_file": "/some/file.xlsx",
        "dependent_files_dir": "/some/dir",
    })
    engine = SyncEngine(str(cfg_file))
    assert engine._central_file == Path("/some/file.xlsx")
    assert engine._dependent_dir == Path("/some/dir")
    assert engine._central_sheet_idx == 0
    assert engine._dep_header_row == 5


# --------------------------------------------------
# _build_lookup
# --------------------------------------------------

def test_build_lookup_creates_mapping(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    central = central_dir / "master.xlsx"
    _create_central_xlsx(central, [
        ("Einwohnerzahl", 50000),
        ("Fallhöhe", 120.5),
        ("Kommunaler Haushalt", 9999),
    ])

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    lookup = engine._build_lookup()

    assert lookup["Einwohnerzahl"] == 50000
    assert lookup["Fallhöhe"] == 120.5
    assert lookup["Kommunaler Haushalt"] == 9999


def test_build_lookup_skips_empty_categories(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    central = central_dir / "master.xlsx"
    _create_central_xlsx(central, [
        ("Valid", 1),
        ("", 2),
        ("  ", 3),
    ])

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    lookup = engine._build_lookup()

    assert lookup == {"Valid": 1}


def test_build_lookup_returns_empty_when_column_missing(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    central = central_dir / "master.xlsx"

    # Create xlsx with wrong column names
    wb = Workbook()
    ws = wb.active
    ws.append(["WrongCol", "AlsoWrong"])
    ws.append(["A", 1])
    wb.save(central)
    wb.close()

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    lookup = engine._build_lookup()

    assert lookup == {}


# --------------------------------------------------
# _find_dependent_files
# --------------------------------------------------

def test_find_dependent_files_returns_xlsx_only(tmp_path: Path):
    dep_dir = tmp_path / "deps"
    dep_dir.mkdir()
    (dep_dir / "A.xlsx").write_bytes(b"x")
    (dep_dir / "B.xlsx").write_bytes(b"x")
    (dep_dir / "notes.txt").write_text("no")

    cfg_file = _write_config(tmp_path, {
        "central_file": str(tmp_path / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    files = engine._find_dependent_files()

    assert sorted(f.name for f in files) == ["A.xlsx", "B.xlsx"]


def test_find_dependent_files_excludes_lock_files(tmp_path: Path):
    dep_dir = tmp_path / "deps"
    dep_dir.mkdir()
    (dep_dir / "Real.xlsx").write_bytes(b"x")
    (dep_dir / "~$Real.xlsx").write_bytes(b"x")

    cfg_file = _write_config(tmp_path, {
        "central_file": str(tmp_path / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    files = engine._find_dependent_files()

    assert [f.name for f in files] == ["Real.xlsx"]


def test_find_dependent_files_returns_empty_for_missing_dir(tmp_path: Path):
    cfg_file = _write_config(tmp_path, {
        "central_file": str(tmp_path / "dummy.xlsx"),
        "dependent_files_dir": str(tmp_path / "nonexistent"),
    })
    engine = SyncEngine(str(cfg_file))

    assert engine._find_dependent_files() == []


# --------------------------------------------------
# _sync_single_file
# --------------------------------------------------

def test_sync_single_file_updates_matching_cells(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    dep_file = dep_dir / "measure.xlsx"
    _create_dependent_xlsx(dep_file, [
        ("Einwohnerzahl", 0),
        ("Fallhöhe", 0),
    ])

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central_dir / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))

    lookup = {"Einwohnerzahl": 50000, "Fallhöhe": 120.5}
    updated = engine._sync_single_file(dep_file, lookup)

    assert updated == 2

    # Verify values were actually written
    from openpyxl import load_workbook
    wb = load_workbook(dep_file)
    ws = wb.worksheets[1]
    assert ws.cell(row=6, column=2).value == 50000
    assert ws.cell(row=7, column=2).value == 120.5
    wb.close()


def test_sync_single_file_skips_unchanged_values(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    dep_file = dep_dir / "measure.xlsx"
    _create_dependent_xlsx(dep_file, [
        ("Einwohnerzahl", 50000),
    ])

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central_dir / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))

    lookup = {"Einwohnerzahl": 50000}
    updated = engine._sync_single_file(dep_file, lookup)

    assert updated == 0


def test_sync_single_file_ignores_unmatched_categories(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    dep_file = dep_dir / "measure.xlsx"
    _create_dependent_xlsx(dep_file, [
        ("Einwohnerzahl", 0),
    ])

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central_dir / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))

    lookup = {"NonExistent": 999}
    updated = engine._sync_single_file(dep_file, lookup)

    assert updated == 0


def test_sync_single_file_skips_if_too_few_sheets(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    dep_file = dep_dir / "single_sheet.xlsx"

    # Create workbook with only one sheet
    wb = Workbook()
    wb.save(dep_file)
    wb.close()

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central_dir / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))

    updated = engine._sync_single_file(dep_file, {"A": 1})

    assert updated == 0


def test_sync_single_file_skips_if_header_columns_missing(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs
    dep_file = dep_dir / "bad_headers.xlsx"

    wb = Workbook()
    wb.active.title = "Overview"
    ws = wb.create_sheet("Daten Generell")
    ws.cell(row=5, column=1, value="WrongHeader")
    ws.cell(row=5, column=2, value="AlsoWrong")
    wb.save(dep_file)
    wb.close()

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central_dir / "dummy.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))

    updated = engine._sync_single_file(dep_file, {"A": 1})

    assert updated == 0


# --------------------------------------------------
# _values_equal
# --------------------------------------------------

@pytest.mark.parametrize(
    "a, b, expected",
    [
        (None, None, True),
        (float("nan"), None, True),
        (None, float("nan"), True),
        (float("nan"), float("nan"), True),
        (None, 1, False),
        (1, None, False),
        (42, 42, True),
        (42, 42.0, True),
        (1.5, 1.5, True),
        (1, 2, False),
        ("abc", "abc", True),
        ("abc", " abc ", True),
        ("abc", "def", False),
        (100, "100", True),
    ],
)
def test_values_equal(a, b, expected):
    assert SyncEngine._values_equal(a, b) == expected


# --------------------------------------------------
# run  (integration)
# --------------------------------------------------

def test_run_syncs_all_dependent_files(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs

    central = central_dir / "master.xlsx"
    _create_central_xlsx(central, [
        ("Einwohnerzahl", 50000),
        ("Fallhöhe", 120.5),
    ])

    _create_dependent_xlsx(dep_dir / "m1.xlsx", [
        ("Einwohnerzahl", 0),
        ("Fallhöhe", 0),
    ])
    _create_dependent_xlsx(dep_dir / "m2.xlsx", [
        ("Einwohnerzahl", 0),
    ])

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    results = engine.run()

    assert results["m1.xlsx"] == 2
    assert results["m2.xlsx"] == 1


def test_run_returns_empty_when_central_file_missing(tmp_path: Path, setup_dirs):
    _, dep_dir = setup_dirs

    cfg_file = _write_config(tmp_path, {
        "central_file": str(tmp_path / "nonexistent.xlsx"),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    results = engine.run()

    assert results == {}


def test_run_returns_empty_when_lookup_is_empty(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs

    central = central_dir / "master.xlsx"
    # Create xlsx with correct headers but no data rows
    wb = Workbook()
    ws = wb.active
    ws.append(["Datenkategorie", "Wert"])
    wb.save(central)
    wb.close()

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    results = engine.run()

    assert results == {}


def test_run_marks_failed_files_with_minus_one(tmp_path: Path, setup_dirs):
    central_dir, dep_dir = setup_dirs

    central = central_dir / "master.xlsx"
    _create_central_xlsx(central, [("Einwohnerzahl", 1)])

    # Write invalid content to create a corrupt xlsx
    bad_file = dep_dir / "corrupt.xlsx"
    bad_file.write_text("not a valid excel file")

    cfg_file = _write_config(tmp_path, {
        "central_file": str(central),
        "dependent_files_dir": str(dep_dir),
    })
    engine = SyncEngine(str(cfg_file))
    results = engine.run()

    assert results["corrupt.xlsx"] == -1
