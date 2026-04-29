import shutil
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.table import Table, TableColumn

# ---------------------------------------------------------------------
# ADJUST THIS IMPORT PATH TO YOUR PROJECT
# e.g.:
# from DatabaseRepositoryLayer.CalculationRepository.preprocessing.structuredRefResolver  import ...
# ---------------------------------------------------------------------
from DatabaseRepositoryLayer.CalculationRepository.preprocessing.structuredRefResolver import (
    resolve_structured_references,
    _is_formula,
    _parse_range,
    _row_range,
    _build_table_map,
    _replace_structured_refs,
    _resolve_bracket,
)


# -------------------------
# Helpers
# -------------------------

def _add_table(ws, name: str, ref: str, columns: list[str]):
    """
    Add an openpyxl Table to worksheet. 'columns' are TableColumn names.
    """
    table = Table(displayName=name, ref=ref)
    table.tableColumns = [TableColumn(id=i + 1, name=col) for i, col in enumerate(columns)]
    ws.add_table(table)
    return table


def _save_wb(wb, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# -------------------------
# Pure helper tests
# -------------------------

def test_is_formula():
    assert _is_formula("=A1+1") is True
    assert _is_formula("A1+1") is False
    assert _is_formula(None) is False
    assert _is_formula(123) is False


def test_parse_range():
    assert _parse_range("E5:G50") == (5, 5, 7, 50)  # E=5, G=7
    assert _parse_range("A1") == (1, 1, 1, 1)


def test_row_range():
    assert _row_range("headers", data_start=6, data_end=50, header_row=5) == (5, 5)
    assert _row_range("totals", data_start=6, data_end=50, header_row=5) == (51, 51)
    assert _row_range("all", data_start=6, data_end=50, header_row=5) == (5, 50)
    assert _row_range("data", data_start=6, data_end=50, header_row=5) == (6, 50)
    assert _row_range(None, data_start=6, data_end=50, header_row=5) == (6, 50)


# -------------------------
# Table map tests
# -------------------------

def test_build_table_map_from_openpyxl_tables():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Table ref: E5:G7 => min_col=5, min_row=5, max_col=7, max_row=7
    _add_table(ws, "Table1", "E5:G7", ["Fallhöhe", "Kommunaler_Haushalt", "Abwassermenge"])

    table_map = _build_table_map(wb)
    assert "table1" in table_map

    t = table_map["table1"]
    assert t["sheet"] == "Sheet1"
    assert t["header_row"] == 5
    assert t["data_start_row"] == 6
    assert t["data_end_row"] == 7
    assert t["min_col"] == 5
    assert t["max_col"] == 7
    # Column letters should match table start col: E,F,G
    assert t["columns"]["fallhöhe"] == "E"
    assert t["columns"]["kommunaler_haushalt"] == "F"
    assert t["columns"]["abwassermenge"] == "G"


# -------------------------
# Bracket resolver tests
# -------------------------

@pytest.fixture
def table_dict_same_sheet():
    # Mimic _build_table_map() entry
    return {
        "sheet": "Sheet1",
        "header_row": 5,
        "data_start_row": 6,
        "data_end_row": 50,
        "min_col": 5,  # E
        "max_col": 7,  # G
        "columns": {
            "fallhöhe": "E",
            "kommunaler_haushalt": "F",
            "abwassermenge": "G",
        },
    }


@pytest.fixture
def table_dict_other_sheet(table_dict_same_sheet):
    t = dict(table_dict_same_sheet)
    t["sheet"] = "OtherSheet"
    return t


def test_resolve_bracket_at_reference_same_sheet(table_dict_same_sheet):
    # Table1[@[Kommunaler_Haushalt]] at current row 20 -> $F$20
    out = _resolve_bracket(
        content="@[Kommunaler_Haushalt]",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=20,
    )
    assert out == "$F$20"

    # Shorthand @Column also supported
    out2 = _resolve_bracket(
        content="@Fallhöhe",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=9,
    )
    assert out2 == "$E$9"


def test_resolve_bracket_adds_sheet_prefix_if_other_sheet(table_dict_other_sheet):
    out = _resolve_bracket(
        content="@[Fallhöhe]",
        table=table_dict_other_sheet,
        current_sheet="Sheet1",
        current_row=10,
    )
    assert out == "'OtherSheet'!$E$10"


def test_resolve_bracket_full_column_defaults_to_data(table_dict_same_sheet):
    # Table1[Fallhöhe] -> full data column range
    out = _resolve_bracket(
        content="Fallhöhe",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out == "$E$6:$E$50"


def test_resolve_bracket_headers_totals_all(table_dict_same_sheet):
    out_h = _resolve_bracket(
        content="[#Headers],[Fallhöhe]",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out_h == "$E$5"

    out_t = _resolve_bracket(
        content="[#Totals],[Fallhöhe]",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out_t == "$E$51"

    out_a = _resolve_bracket(
        content="[#All],[Fallhöhe]",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out_a == "$E$5:$E$50"


def test_resolve_bracket_col_range(table_dict_same_sheet):
    # [Fallhöhe]:[Abwassermenge] => E..G over data rows
    out = _resolve_bracket(
        content="[Fallhöhe]:[Abwassermenge]",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out == "$E$6:$G$50"


def test_resolve_bracket_empty_content_entire_data_range(table_dict_same_sheet):
    # Table1[] -> whole table section (defaults to data range, all columns)
    out = _resolve_bracket(
        content="",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out == "$E$6:$G$50"


def test_resolve_bracket_unknown_column_returns_ref_error(table_dict_same_sheet):
    out = _resolve_bracket(
        content="[@[DoesNotExist]]",
        table=table_dict_same_sheet,
        current_sheet="Sheet1",
        current_row=1,
    )
    assert out == "#REF!"


# -------------------------
# Formula replacement tests
# -------------------------

def test_replace_structured_refs_rewrites_multiple_refs(table_dict_same_sheet):
    table_map = {"table1": table_dict_same_sheet}

    formula = "=SUM(Table1[[#Data],[Fallhöhe]])+Table1[@[Kommunaler_Haushalt]]"
    out = _replace_structured_refs(
        formula=formula,
        table_map=table_map,
        current_sheet_name="Sheet1",
        current_row=20,
    )

    assert out == "=SUM($E$6:$E$50)+$F$20"


def test_replace_structured_refs_unknown_table_left_unchanged(table_dict_same_sheet):
    table_map = {"table1": table_dict_same_sheet}
    formula = "=SUM(NotATable[[#Data],[Fallhöhe]])"
    out = _replace_structured_refs(
        formula=formula,
        table_map=table_map,
        current_sheet_name="Sheet1",
        current_row=1,
    )
    assert out == formula


# -------------------------
# End-to-end resolve_structured_references tests
# -------------------------

def test_resolve_structured_references_returns_original_if_no_tables(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].value = "=1+2"

    p = tmp_path / "no_tables.xlsx"
    _save_wb(wb, p)

    out = resolve_structured_references(p)
    assert out == p


def test_resolve_structured_references_returns_original_if_no_structured_refs(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _add_table(ws, "Table1", "E5:G7", ["Fallhöhe", "Kommunaler_Haushalt", "Abwassermenge"])

    ws["A1"].value = "=1+2"  # formula but no structured ref

    p = tmp_path / "no_structured_refs.xlsx"
    _save_wb(wb, p)

    out = resolve_structured_references(p)
    assert out == p


def test_resolve_structured_references_creates_temp_file_when_modified(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Table E5:G7 -> data rows 6..7
    _add_table(ws, "Table1", "E5:G7", ["Fallhöhe", "Kommunaler_Haushalt", "Abwassermenge"])

    ws["A1"].value = "=SUM(Table1[[#Data],[Fallhöhe]])"  # structured ref present

    p = tmp_path / "structured.xlsx"
    _save_wb(wb, p)

    out = resolve_structured_references(p)

    try:
        assert out != p
        assert out.exists()

        # verify formula actually got rewritten
        wb2 = openpyxl.load_workbook(out, data_only=False)
        ws2 = wb2["Sheet1"]
        assert ws2["A1"].value == "=SUM($E$6:$E$7)"
    finally:
        # cleanup temp dir created by resolver
        if out != p:
            shutil.rmtree(out.parent, ignore_errors=True)
