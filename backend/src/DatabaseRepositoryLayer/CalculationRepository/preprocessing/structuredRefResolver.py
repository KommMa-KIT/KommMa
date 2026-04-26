"""
StructuredRefResolver
=====================
Pre-processing step that rewrites Excel **structured table references**
(e.g. ``Table1[[#Data],[Fallhoehe]]``) into plain absolute cell addresses
(e.g. ``$E$6:$E$50``) that the ``formulas`` library can parse.

Business Purpose
----------------
Many measure workbooks use Excel Tables whose formulae rely on structured
references.  The ``formulas`` parser does not support this syntax, so this
module bridges the gap by performing a one-time rewrite before compilation.

The original file is **never** modified; a temporary copy is created and
returned to the caller, who must delete it when done.

Scalability
-----------
If future versions of the ``formulas`` library add native structured-reference
support this entire module can be removed without touching any other code --
just skip the pre-processing step in ``FormulaEngine._build_model``.

Thread Safety
-------------
The public function ``resolve_structured_references`` creates its own
temporary directory and ``openpyxl.Workbook``; it is safe for concurrent use.
"""

from __future__ import annotations

import re
import tempfile
import logging
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def resolve_structured_references(xlsx_path: Path) -> Path:
    """
    Rewrite all structured table references in *xlsx_path* to absolute
    cell addresses.

    If structured references are found, a **temporary** copy of the
    workbook with resolved references is returned.  If none are found,
    the original path is returned unchanged (no temp file created).

    Caller contract
        The caller **must** delete the temporary directory when finished::

            clean = resolve_structured_references(p)
            try:
                # ... use clean ...
            finally:
                if clean != p:
                    shutil.rmtree(clean.parent, ignore_errors=True)

    @param xlsx_path: Path to the original ``.xlsx`` workbook.

    @return: Path to a (possibly temporary) workbook with all structured
             references replaced.  May equal *xlsx_path* if no rewriting
             was needed.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    table_map = _build_table_map(wb)

    if not table_map:
        logger.debug("No tables found, skipping structured-reference resolver")
        return xlsx_path

    modified = False

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not _is_formula(cell.value):
                    continue

                new_formula = _replace_structured_refs(
                    formula=cell.value,
                    table_map=table_map,
                    current_sheet_name=sheet.title,
                    current_row=cell.row,
                )

                if new_formula != cell.value:
                    cell.value = new_formula
                    modified = True
                    logger.debug("Resolved structured reference in formula cell")

    if not modified:
        logger.debug("No structured references found in formulas")
        return xlsx_path

    tmpdir = Path(tempfile.mkdtemp())
    temp_path = tmpdir / xlsx_path.name
    wb.save(temp_path)
    logger.info("Structured references resolved into temporary workbook")
    return temp_path


# ---------------------------------------------------------------------------
# Audit helper (run once to see every structured ref across all files)
# ---------------------------------------------------------------------------

def audit_structured_refs(xlsx_dir: Path) -> None:
    """
    Print every formula cell that contains a potential structured reference.

    This is a **development-only** diagnostic tool.  Use it to verify
    coverage before switching from xlwings to the ``formulas``-based engine.

    @param xlsx_dir: Directory of ``.xlsx`` files to scan.
    """
    pattern = re.compile(r'[A-Za-z_]\w*\[', re.IGNORECASE)

    for xlsx_file in sorted(xlsx_dir.glob("*.xlsx")):
        if xlsx_file.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=False)
        except Exception as e:
            logger.warning("Could not open workbook during structured-ref audit")
            continue

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if _is_formula(cell.value) and pattern.search(cell.value):
                        logger.info("Structured reference found in workbook formula cell")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _is_formula(value) -> bool:
    """Return ``True`` if *value* looks like an Excel formula (starts with ``=``)."""
    return isinstance(value, str) and value.startswith("=")


def _build_table_map(wb: openpyxl.Workbook) -> Dict[str, dict]:
    """
    Index every ``ListObject`` (Excel Table) in the workbook.

    The returned dict is keyed by **lower-case** table name and contains
    the row/column boundaries plus a column-name -> letter mapping needed
    by ``_resolve_bracket`` to translate structured references.

    @param wb: An ``openpyxl.Workbook`` opened with ``data_only=False``.

    @return: ``{table_name_lower: {"sheet", "header_row", "data_start_row",
             "data_end_row", "min_col", "max_col", "columns": {...}}, ...}``
    """
    table_map: Dict[str, dict] = {}

    for sheet in wb.worksheets:
        for table in sheet.tables.values():
            min_col, min_row, max_col, max_row = _parse_range(table.ref)

            col_map: Dict[str, str] = {}
            for i, tcol in enumerate(table.tableColumns):
                letter = _col_idx_to_letter(min_col + i)   # min_col is 1-based
                col_map[tcol.name.lower().strip()] = letter

            table_map[table.name.lower()] = {
                "sheet":          sheet.title,
                "header_row":     min_row,
                "data_start_row": min_row + 1,
                "data_end_row":   max_row,
                "min_col":        min_col,
                "max_col":        max_col,
                "columns":        col_map,
            }

    return table_map


# Matches  TableName[...]  including nested brackets like  T[[#Data],[Col]]
_STRUCT_REF_RE = re.compile(
    r'(?<![A-Za-z0-9_])'           # not preceded by identifier chars
    r'([A-Za-z_][A-Za-z0-9_. ]*)'  # table name (spaces allowed in Excel names)
    r'\[('                          # opening bracket — start capture
    r'(?:[^\[\]]|\[[^\[\]]*\])*'    # content: non-brackets OR [inner]
    r')\]',                         # closing bracket
    re.IGNORECASE,
)


def _replace_structured_refs(
    formula: str,
    table_map: Dict[str, dict],
    current_sheet_name: str,
    current_row: int,
) -> str:
    """
    Replace all structured table references in a single formula string.

    Uses ``_STRUCT_REF_RE`` to find ``TableName[...]`` patterns and
    delegates resolution to ``_resolve_bracket``.

    @param formula:            The formula string (including leading ``=``).
    @param table_map:          Table index built by ``_build_table_map``.
    @param current_sheet_name: Worksheet title where the formula resides.
    @param current_row:        1-based row of the cell (needed for ``@``
                               implicit-intersection references).

    @return: Formula with all structured references replaced by absolute
             cell addresses.  Unchanged if no references are found.
    """

    def replacer(m: re.Match) -> str:
        table_name = m.group(1).lower().strip()
        bracket_content = m.group(2)

        if table_name not in table_map:
            return m.group(0)           # not a table — leave as-is

        table = table_map[table_name]
        resolved = _resolve_bracket(
            bracket_content,
            table,
            current_sheet_name,
            current_row,
        )
        logger.debug(f"    {m.group(0)} → {resolved}")
        return resolved

    return _STRUCT_REF_RE.sub(replacer, formula)


def _resolve_bracket(
    content: str,
    table: dict,
    current_sheet: str,
    current_row: int,
) -> str:
    """
    Translate the bracket content of a single structured reference into an
    absolute cell or range address.

    Supported patterns (content is everything *inside* the outer brackets):

        @[Column]           -- current-row cell (implicit intersection)
        @Column             -- shorthand for the above
        [Column]            -- full data column  (= [#Data],[Column])
        [#Data],[Column]    -- explicit data column
        [#Headers],[Column] -- header cell
        [#Totals],[Column]  -- totals row cell
        [#All],[Column]     -- header + data + totals
        [Col1]:[Col2]       -- multi-column range (data rows)
        (empty)             -- entire table data range

    @param content:       Text inside the outermost brackets.
    @param table:         Table descriptor from ``_build_table_map``.
    @param current_sheet: Worksheet title of the formula cell.
    @param current_row:   1-based row of the formula cell.

    @return: Absolute cell/range reference, or ``"#REF!"`` if the column
             name cannot be resolved.
    """
    sheet_prefix = (
        f"'{table['sheet']}'!" if table["sheet"] != current_sheet else ""
    )

    data_start = table["data_start_row"]
    data_end   = table["data_end_row"]
    header_row = table["header_row"]
    
    # ── Column range pattern: [Col1]:[Col2] ───────────────────────────────
    # Example: [Fallhöhe]:[Abwassermenge]
    m_range = re.match(r'^\s*\[([^\[\]]+)\]\s*:\s*\[([^\[\]]+)\]\s*$', content)
    if m_range:
        col_name1 = m_range.group(1).strip().lower()
        col_name2 = m_range.group(2).strip().lower()

        col_letter1 = table["columns"].get(col_name1)
        col_letter2 = table["columns"].get(col_name2)

        if col_letter1 is None or col_letter2 is None:
            logger.warning("Unknown column in structured range reference")
            return "#REF!"

        # Range is over data rows by default (consistent with docstring)
        row_start, row_end = _row_range(None, data_start, data_end, header_row)
        return f"{sheet_prefix}${col_letter1}${row_start}:${col_letter2}${row_end}"


    # ── @ implicit-intersection (current row) ──────────────────────────────
    at_match = re.match(r'^@\[?([^\[\]]+)\]?$', content.strip())
    if at_match:
        col_name = at_match.group(1).lower().strip()
        col_letter = table["columns"].get(col_name)
        if col_letter:
            return f"{sheet_prefix}${col_letter}${current_row}"
        logger.warning("Unknown column in structured @ reference")
        return "#REF!"

    # ── Parse [#Specifier] and [Column] parts ─────────────────────────────
    tokens = re.findall(r'\[([^\[\]]+)\]', content)

    specifier: Optional[str] = None
    col_name:  Optional[str] = None
    col_name2: Optional[str] = None   # for [Col1]:[Col2]

    for tok in tokens:
        tok_s = tok.strip()
        if tok_s.startswith("#"):
            specifier = tok_s[1:].lower()   # 'data', 'headers', 'totals', 'all'
        elif ":" in tok_s:
            parts = tok_s.split(":", 1)
            col_name  = parts[0].strip().strip("[]").lower()
            col_name2 = parts[1].strip().strip("[]").lower()
        else:
            col_name = tok_s.lower()

    # Bare content with no brackets:
    # - empty => entire table section (data/all etc.)
    # - non-empty => column name
    if not tokens:
        if content.strip() == "":
            col_name = None
        else:
            col_name = content.strip().lower()


    row_start, row_end = _row_range(specifier, data_start, data_end, header_row)

    # ── No column → whole table (or section) ──────────────────────────────
    if col_name is None:
        c_start = _col_idx_to_letter(table["min_col"])
        c_end   = _col_idx_to_letter(table["max_col"])
        return f"{sheet_prefix}${c_start}${row_start}:${c_end}${row_end}"

    col_letter = table["columns"].get(col_name)
    if col_letter is None:
        logger.warning("Unknown structured-reference column")
        return "#REF!"

    # ── Column range [Col1]:[Col2] ─────────────────────────────────────────
    if col_name2:
        col_letter2 = table["columns"].get(col_name2)
        if col_letter2 is None:
            logger.warning("Unknown structured-reference column")
            return "#REF!"
        return f"{sheet_prefix}${col_letter}${row_start}:${col_letter2}${row_end}"

    # ── Single column ──────────────────────────────────────────────────────
    if row_start == row_end:
        return f"{sheet_prefix}${col_letter}${row_start}"
    return f"{sheet_prefix}${col_letter}${row_start}:${col_letter}${row_end}"


def _row_range(
    specifier: Optional[str],
    data_start: int,
    data_end: int,
    header_row: int,
) -> Tuple[int, int]:
    """
    Map a structured-reference specifier to the corresponding row range.

    @param specifier:  One of ``"headers"``, ``"totals"``, ``"all"``,
                       ``"data"``, or ``None`` (treated as ``"data"``).
    @param data_start: First data row (1-based).
    @param data_end:   Last data row (1-based).
    @param header_row: Header row (1-based).

    @return: ``(start_row, end_row)`` tuple (both 1-based, inclusive).
    """
    if specifier == "headers":
        return header_row, header_row
    if specifier == "totals":
        totals_row = data_end + 1
        return totals_row, totals_row
    if specifier == "all":
        return header_row, data_end
    # 'data' or None → just data rows
    return data_start, data_end


def _parse_range(ref: str) -> Tuple[int, int, int, int]:
    """
    Parse an Excel range string into numeric boundaries.

    @param ref: Range string, e.g. ``"E5:G50"`` or a single cell ``"E5"``.

    @return: ``(min_col_1based, min_row, max_col_1based, max_row)``.
    """
    if ":" in ref:
        left, right = ref.split(":", 1)
    else:
        left = right = ref

    left_col  = "".join(c for c in left  if c.isalpha())
    left_row  = int("".join(c for c in left  if c.isdigit()))
    right_col = "".join(c for c in right if c.isalpha())
    right_row = int("".join(c for c in right if c.isdigit()))

    return (
        column_index_from_string(left_col),
        left_row,
        column_index_from_string(right_col),
        right_row,
    )


def _col_idx_to_letter(idx: int) -> str:
    """
    Convert a 1-based column index to an Excel column letter.

    @param idx: 1-based index (1 -> A, 26 -> Z, 27 -> AA).

    @return: Upper-case column letter string.
    """
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result