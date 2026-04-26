"""
Sync Engine
===========
Core logic for propagating ``Wert`` values from the central mastertable into
every dependent calculation Excel file.

Design decisions
----------------
* **openpyxl** is used for all writes so that cell formatting, formulas on
  other sheets, and workbook structure are fully preserved.
* **pandas** is used only for fast in-memory matching by ``Datenkategorie``.
* Only the ``Wert`` column on **Sheet index 1** (row 5 header) of each
  dependent file is ever modified — no other sheet or column is touched.
* The engine builds a ``{Datenkategorie: Wert}`` lookup from the central file
  once, then iterates over dependent files, making it O(n·m) in the worst
  case but with a constant-time lookup per row.
* Writes are atomic per file: the workbook is saved only after all matching
  rows have been updated, so a crash mid-run cannot leave a half-written file.

Thread Safety
-------------
A single ``SyncEngine`` instance is **not** thread-safe.  The watcher module
serialises calls via a debounce timer, so concurrent invocations do not occur
in normal operation.
"""

import json
import logging
import os
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class SyncEngine:
    """
    Reads the central mastertable and pushes updated ``Wert`` values into
    every dependent calculation sheet found in the configured directory.
    """

    def __init__(self, config_path: str):
        with open(config_path, "r", encoding="utf-8") as fh:
            self._cfg: dict[str, Any] = json.load(fh)

        self._central_file = Path(self._cfg["central_file"])
        self._dependent_dir = Path(self._cfg["dependent_files_dir"])
        self._central_sheet_idx: int = self._cfg["central_sheet_index"]
        self._central_header_row: int = self._cfg["central_header_row"]
        self._central_match_col: str = self._cfg["central_match_column"]
        self._central_value_col: str = self._cfg["central_value_column"]
        self._dep_sheet_idx: int = self._cfg["dependent_sheet_index"]
        self._dep_header_row: int = self._cfg["dependent_header_row"]
        self._dep_match_col: str = self._cfg["dependent_match_column"]
        self._dep_value_col: str = self._cfg["dependent_value_column"]

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def run(self) -> dict[str, int]:
        """
        Execute a full sync cycle.

        @return: ``{filename: number_of_updated_cells}`` for every dependent
                 file that was touched.  Files with zero matches are included
                 with value ``0``.
        """
        if not self._central_file.is_file():
            logger.error("Central file not found")
            return {}

        lookup = self._build_lookup()
        if not lookup:
            logger.warning("Central file produced an empty lookup — nothing to sync.")
            return {}

        logger.info(
            "Loaded %d categories from central file",
            len(lookup),
        )

        results: dict[str, int] = {}

        for xlsx in self._find_dependent_files():
            try:
                updated = self._sync_single_file(xlsx, lookup)
                results[xlsx.name] = updated
                if updated:
                    logger.info("Updated %d cells in a dependent workbook", updated)
                else:
                    logger.debug("No matching categories in a dependent workbook")
            except Exception:
                logger.exception("Failed to sync a dependent workbook")
                results[xlsx.name] = -1

        total = sum(v for v in results.values() if v > 0)
        logger.info("Sync complete — %d cells updated across %d files.", total, len(results))
        return results

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _build_lookup(self) -> dict[str, Any]:
        """
        Build a ``{Datenkategorie: Wert}`` mapping from the central file.

        Uses pandas for fast column-based access.  The header row in the
        central file is row 1 (0-indexed row 0 for pandas).
        """
        df = pd.read_excel(
            self._central_file,
            sheet_name=self._central_sheet_idx,
            header=self._central_header_row - 1,  # pandas 0-based
            engine="openpyxl",
        )

        if self._central_match_col not in df.columns:
                logger.error("Configured central match column not found")
            return {}

        if self._central_value_col not in df.columns:
                logger.error("Configured central value column not found")
            return {}

        lookup: dict[str, Any] = {}
        for _, row in df.iterrows():
            cat = row[self._central_match_col]
            if pd.isna(cat) or str(cat).strip() == "":
                continue
            lookup[str(cat).strip()] = row[self._central_value_col]

        return lookup

    def _find_dependent_files(self) -> list[Path]:
        """
        Return all ``.xlsx`` files in the dependent directory, excluding
        Excel lock files (``~$`` prefix).
        """
        if not self._dependent_dir.is_dir():
            logger.error("Dependent directory not found")
            return []

        return sorted(
            p
            for p in self._dependent_dir.iterdir()
            if p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
        )

    def _sync_single_file(self, xlsx_path: Path, lookup: dict[str, Any]) -> int:
        """
        Open *xlsx_path* with openpyxl, update matching ``Wert`` cells on
        Sheet index 1, and save the workbook.  Returns the number of cells
        actually written.

        Only cells whose current value differs from the lookup value are
        overwritten, minimising unnecessary disk I/O and preserving the
        file's modification timestamp when nothing changes.
        """
        wb = load_workbook(xlsx_path)

        if len(wb.worksheets) <= self._dep_sheet_idx:
            logger.warning(
                "Workbook has only %d sheet(s) — expected at least %d. Skipping.",
                len(wb.worksheets),
                self._dep_sheet_idx + 1,
            )
            wb.close()
            return 0

        ws = wb.worksheets[self._dep_sheet_idx]

        # Locate the match and value columns by scanning the header row.
        match_col_idx: int | None = None
        value_col_idx: int | None = None

        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=self._dep_header_row, column=col).value
            if header is None:
                continue
            header_str = str(header).strip()
            if header_str == self._dep_match_col:
                match_col_idx = col
            elif header_str == self._dep_value_col:
                value_col_idx = col

        if match_col_idx is None or value_col_idx is None:
            logger.warning(
                "Workbook: could not locate '%s' (col %s) or '%s' (col %s) "
                "in header row %d. Skipping.",
                self._dep_match_col,
                match_col_idx,
                self._dep_value_col,
                value_col_idx,
                self._dep_header_row,
            )
            wb.close()
            return 0

        updated = 0
        row = self._dep_header_row + 1  # first data row

        while row <= ws.max_row:
            cat_cell = ws.cell(row=row, column=match_col_idx).value
            if cat_cell is None or str(cat_cell).strip() == "":
                break  # end of data section

            cat_key = str(cat_cell).strip()
            if cat_key in lookup:
                new_value = lookup[cat_key]
                current = ws.cell(row=row, column=value_col_idx).value

                # Only write when the value actually changed.
                if not self._values_equal(current, new_value):
                    ws.cell(row=row, column=value_col_idx).value = new_value
                    updated += 1

            row += 1

        if updated > 0:
            wb.save(xlsx_path)
            logger.debug("Saved workbook with %d change(s).", updated)

        wb.close()
        return updated

    @staticmethod
    def _values_equal(a: Any, b: Any) -> bool:
        """
        Compare two cell values, treating NaN/None as equal and performing
        numeric comparison when both sides are numbers.
        """
        a_empty = a is None or (isinstance(a, float) and pd.isna(a))
        b_empty = b is None or (isinstance(b, float) and pd.isna(b))

        if a_empty and b_empty:
            return True
        if a_empty != b_empty:
            return False

        # Numeric comparison avoids float representation mismatches.
        try:
            return float(a) == float(b)
        except (ValueError, TypeError):
            return str(a).strip() == str(b).strip()
